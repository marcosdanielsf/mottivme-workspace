#!/usr/bin/env python3
"""
GHL Agency AI - Cost Analysis Spreadsheet Generator (with Browserbase)
Generates comprehensive cost breakdown for different usage scenarios
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_cost_analysis():
    wb = Workbook()
    wb.remove(wb.active)
    
    # ===== SHEET 1: Monthly Costs Summary =====
    ws_summary = wb.create_sheet("Monthly Costs Summary")
    
    summary_data = {
        'Service Provider': [
            'GoHighLevel',
            'Notion',
            'Slack',
            'Neon Database',
            'Browserbase',
            'Twilio (SMS/Voice)',
            'Twilio (WhatsApp)',
            'Google Gemini API',
            'Google Workspace',
            'Vercel Hosting',
            'GitHub',
            '',
            'SUBTOTAL (Platform Services)',
            '',
            'Development Costs (One-time)',
            'Monthly Maintenance',
            '',
            'TOTAL MONTHLY (Operational)',
            'TOTAL FIRST MONTH (w/ Development)'
        ],
        'Starter Tier\n(1-5 clients)': [
            '$297',
            '$40',
            '$22',
            '$19',
            '$20',
            '$15',
            '$5',
            '$18',
            '$12',
            '$20',
            '$0',
            '',
            '$468',
            '',
            '$18,000',
            '$750',
            '',
            '$1,218',
            '$19,218'
        ],
        'Growth Tier\n(5-20 clients)': [
            '$297',
            '$60',
            '$36',
            '$69',
            '$99',
            '$40',
            '$15',
            '$45',
            '$24',
            '$20',
            '$0',
            '',
            '$705',
            '',
            '$18,000',
            '$1,500',
            '',
            '$2,205',
            '$20,205'
        ],
        'Enterprise Tier\n(20+ clients)': [
            '$497',
            '$100',
            '$72',
            '$700',
            '$500',
            '$100',
            '$40',
            '$120',
            '$36',
            '$20',
            '$0',
            '',
            '$2,185',
            '',
            '$18,000',
            '$3,000',
            '',
            '$5,185',
            '$23,185'
        ],
        'Notes': [
            'Agency Unlimited/Pro plan with API access',
            'Business plan for 2-5 users with AI',
            'Pro plan for 3-10 team members',
            'Launch/Scale/Business plan based on tier',
            'Developer/Startup/Scale plan (cloud browsers)',
            'Based on estimated message volume',
            'WhatsApp conversation fees',
            'Gemini Flash for high-volume operations',
            'Business Standard for storage',
            'Pro plan for production deployment',
            'Free for public repositories',
            '',
            'Sum of all platform service costs',
            '',
            'Custom integration development (10 services)',
            'Ongoing support and maintenance',
            '',
            'Monthly recurring operational costs',
            'Includes one-time development investment'
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif r_idx in [14, 19, 20]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            elif r_idx in [13, 15, 17]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:
                cell.alignment = Alignment(horizontal="left" if c_idx == 1 or c_idx == 5 else "center", vertical="center", wrap_text=True)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 18
    ws_summary.column_dimensions['E'].width = 45
    ws_summary.row_dimensions[1].height = 30
    
    # Save workbook
    output_file = '/home/ubuntu/ghl-agency-ai/GHL-Agency-AI-Cost-Analysis-Updated.xlsx'
    wb.save(output_file)
    print(f"Updated cost analysis spreadsheet created: {output_file}")
    return output_file

if __name__ == "__main__":
    create_cost_analysis()

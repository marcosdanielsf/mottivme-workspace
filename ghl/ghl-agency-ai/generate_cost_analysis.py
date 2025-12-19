#!/usr/bin/env python3
"""
GHL Agency AI - Cost Analysis Spreadsheet Generator
Generates comprehensive cost breakdown for different usage scenarios
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_cost_analysis():
    """Generate comprehensive cost analysis spreadsheet"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ===== SHEET 1: Monthly Costs Summary =====
    ws_summary = wb.create_sheet("Monthly Costs Summary")
    
    summary_data = {
        'Service Provider': [
            'GoHighLevel',
            'Notion',
            'Slack',
            'Neon Database',
            'Twilio (SMS/Voice)',
            'Twilio (WhatsApp)',
            'Google Gemini API',
            'Google Workspace',
            'Vercel Hosting',
            'GitHub',
            '',
            'SUBTOTAL (Platform Services)',
            '',
            'Development Costs (One-time)',
            'Monthly Maintenance',
            '',
            'TOTAL MONTHLY (Operational)',
            'TOTAL FIRST MONTH (w/ Development)'
        ],
        'Starter Tier\n(1-5 clients)': [
            '$297',
            '$40',
            '$22',
            '$19',
            '$15',
            '$5',
            '$18',
            '$12',
            '$20',
            '$0',
            '',
            '$448',
            '',
            '$15,000',
            '$750',
            '',
            '$1,198',
            '$16,198'
        ],
        'Growth Tier\n(5-20 clients)': [
            '$297',
            '$60',
            '$36',
            '$69',
            '$40',
            '$15',
            '$45',
            '$24',
            '$20',
            '$0',
            '',
            '$606',
            '',
            '$15,000',
            '$1,500',
            '',
            '$2,106',
            '$17,106'
        ],
        'Enterprise Tier\n(20+ clients)': [
            '$497',
            '$100',
            '$72',
            '$700',
            '$100',
            '$40',
            '$120',
            '$36',
            '$20',
            '$0',
            '',
            '$1,685',
            '',
            '$15,000',
            '$3,000',
            '',
            '$4,685',
            '$19,685'
        ],
        'Notes': [
            'Agency Unlimited plan with API access',
            'Business plan for 2-5 users with AI',
            'Pro plan for 3-10 team members',
            'Launch/Scale/Business plan based on tier',
            'Based on estimated message volume',
            'WhatsApp conversation fees',
            'Gemini Flash for high-volume operations',
            'Business Standard for storage',
            'Pro plan for production deployment',
            'Free for public repositories',
            '',
            'Sum of all platform service costs',
            '',
            'Custom integration development (9 services)',
            'Ongoing support and maintenance',
            '',
            'Monthly recurring operational costs',
            'Includes one-time development investment'
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            
            # Header row styling
            if r_idx == 1:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Subtotal and Total rows
            elif r_idx in [13, 18, 19]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Empty separator rows
            elif r_idx in [12, 14, 16]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Regular cells
            else:
                cell.alignment = Alignment(horizontal="left" if c_idx == 1 or c_idx == 5 else "center", vertical="center", wrap_text=True)
    
    # Set column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 18
    ws_summary.column_dimensions['E'].width = 40
    
    # Set row heights
    ws_summary.row_dimensions[1].height = 30
    
    # ===== SHEET 2: Detailed Usage Scenarios =====
    ws_usage = wb.create_sheet("Usage Scenarios")
    
    usage_data = {
        'Metric': [
            'Number of Clients',
            'Active AI Agents',
            'GHL Sub-accounts',
            'Team Members',
            'Monthly SMS Messages',
            'Monthly Voice Minutes',
            'Monthly WhatsApp Messages',
            'Gemini API Requests',
            'Notion Pages Accessed',
            'Drive Files Accessed',
            'Slack Notifications',
            'Database Storage (GB)',
            'Bandwidth (GB)',
            '',
            'Estimated Monthly Cost',
            'Cost per Client',
            'Cost per Agent Hour'
        ],
        'Starter': [
            '1-5',
            '2-5',
            '5',
            '2-3',
            '500',
            '200',
            '200',
            '5,000',
            '50',
            '100',
            '500',
            '2',
            '50',
            '',
            '$1,198',
            '$240',
            '$8'
        ],
        'Growth': [
            '5-20',
            '5-15',
            '20',
            '5-8',
            '2,000',
            '800',
            '800',
            '15,000',
            '200',
            '500',
            '2,000',
            '10',
            '200',
            '',
            '$2,106',
            '$105',
            '$5'
        ],
        'Enterprise': [
            '20-100',
            '20-50',
            'Unlimited',
            '10-20',
            '10,000',
            '3,000',
            '3,000',
            '50,000',
            '1,000',
            '2,000',
            '10,000',
            '100',
            '1,000',
            '',
            '$4,685',
            '$47',
            '$3'
        ]
    }
    
    df_usage = pd.DataFrame(usage_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_usage, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_usage.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif r_idx in [15, 16, 17, 18]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif r_idx == 14:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:
                cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
    
    ws_usage.column_dimensions['A'].width = 30
    ws_usage.column_dimensions['B'].width = 15
    ws_usage.column_dimensions['C'].width = 15
    ws_usage.column_dimensions['D'].width = 15
    
    # ===== SHEET 3: ROI Calculator =====
    ws_roi = wb.create_sheet("ROI Calculator")
    
    roi_data = {
        'Cost Category': [
            'CURRENT COSTS (Manual Operations)',
            'VA Salary (Full-time)',
            'Manager Salary (Part-time)',
            'Software Subscriptions',
            'Training & Onboarding',
            'Error Correction & Rework',
            '',
            'TOTAL MONTHLY (Manual)',
            '',
            'PROPOSED COSTS (AI Automation)',
            'Platform Subscription',
            'Development (Amortized over 12 months)',
            'Monthly Maintenance',
            '',
            'TOTAL MONTHLY (Automated)',
            '',
            'MONTHLY SAVINGS',
            'ANNUAL SAVINGS',
            'ROI Timeline (Months)',
            'Efficiency Gain (%)'
        ],
        'Starter Tier': [
            '',
            '$3,000',
            '$2,000',
            '$500',
            '$300',
            '$500',
            '',
            '$6,300',
            '',
            '',
            '$448',
            '$1,250',
            '$750',
            '',
            '$2,448',
            '',
            '$3,852',
            '$46,224',
            '3.9',
            '61%'
        ],
        'Growth Tier': [
            '',
            '$6,000',
            '$4,000',
            '$1,000',
            '$600',
            '$1,000',
            '',
            '$12,600',
            '',
            '',
            '$606',
            '$1,250',
            '$1,500',
            '',
            '$3,356',
            '',
            '$9,244',
            '$110,928',
            '1.6',
            '73%'
        ],
        'Enterprise Tier': [
            '',
            '$15,000',
            '$8,000',
            '$2,500',
            '$1,500',
            '$3,000',
            '',
            '$30,000',
            '',
            '',
            '$1,685',
            '$1,250',
            '$3,000',
            '',
            '$5,935',
            '',
            '$24,065',
            '$288,780',
            '0.6',
            '80%'
        ]
    }
    
    df_roi = pd.DataFrame(roi_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_roi, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_roi.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif r_idx in [2, 10]:
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            elif r_idx in [8, 15, 17, 18, 19, 20]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif r_idx in [7, 9, 14, 16]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:
                cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
    
    ws_roi.column_dimensions['A'].width = 35
    ws_roi.column_dimensions['B'].width = 18
    ws_roi.column_dimensions['C'].width = 18
    ws_roi.column_dimensions['D'].width = 18
    
    # ===== SHEET 4: Integration Development Costs =====
    ws_dev = wb.create_sheet("Development Breakdown")
    
    dev_data = {
        'Integration': [
            'GoHighLevel API',
            'Notion API',
            'Slack Webhooks',
            'Neon Database Setup',
            'Twilio SMS/Voice',
            'WhatsApp Business API',
            'Google Gemini AI',
            'Google Drive API',
            'Vercel Deployment',
            '',
            'Core Application',
            'Dashboard UI',
            'Agent Orchestration',
            'Browser Automation',
            'Authentication System',
            'Testing & QA',
            '',
            'TOTAL DEVELOPMENT COST'
        ],
        'Hours': [
            '8-12',
            '6-8',
            '4-6',
            '6-8',
            '8-10',
            '6-8',
            '10-12',
            '6-8',
            '4-6',
            '',
            '',
            '20-30',
            '30-40',
            '15-20',
            '10-15',
            '20-30',
            '',
            '153-213'
        ],
        'Rate ($/hr)': [
            '$100',
            '$100',
            '$75',
            '$100',
            '$100',
            '$100',
            '$125',
            '$100',
            '$75',
            '',
            '',
            '$100',
            '$125',
            '$100',
            '$100',
            '$75',
            '',
            ''
        ],
        'Cost Range': [
            '$800-$1,200',
            '$600-$800',
            '$300-$450',
            '$600-$800',
            '$800-$1,000',
            '$600-$800',
            '$1,250-$1,500',
            '$600-$800',
            '$300-$450',
            '',
            '',
            '$2,000-$3,000',
            '$3,750-$5,000',
            '$1,500-$2,000',
            '$1,000-$1,500',
            '$1,500-$2,250',
            '',
            '$15,600-$21,550'
        ],
        'Complexity': [
            'Medium',
            'Low',
            'Low',
            'Medium',
            'Medium',
            'Medium',
            'High',
            'Medium',
            'Low',
            '',
            '',
            'High',
            'High',
            'Medium',
            'Medium',
            'Medium',
            '',
            ''
        ]
    }
    
    df_dev = pd.DataFrame(dev_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_dev, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_dev.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif r_idx in [11, 18]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif r_idx in [10, 17]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:
                cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
    
    ws_dev.column_dimensions['A'].width = 30
    ws_dev.column_dimensions['B'].width = 12
    ws_dev.column_dimensions['C'].width = 12
    ws_dev.column_dimensions['D'].width = 18
    ws_dev.column_dimensions['E'].width = 12
    
    # Save workbook
    output_file = '/home/ubuntu/ghl-agency-ai/GHL-Agency-AI-Cost-Analysis.xlsx'
    wb.save(output_file)
    print(f"Cost analysis spreadsheet created: {output_file}")
    return output_file

if __name__ == "__main__":
    create_cost_analysis()
